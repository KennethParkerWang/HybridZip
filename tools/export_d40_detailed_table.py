#!/usr/bin/env python3
"""Export the validated D40 ledger into review-friendly XLSX and Markdown."""

from __future__ import annotations

import csv
import hashlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
D40 = ROOT / "results" / "analysis" / "r2-silesia-32k-currenthash-20260821-d40"
XLSX = D40 / "D40_detailed_table.xlsx"
MARKDOWN = D40 / "D40_detailed_table.md"


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream, delimiter="\t"))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def as_int(row: dict[str, str], key: str) -> int:
    return int(float(row[key]))


def as_float(row: dict[str, str], key: str) -> float:
    return float(row[key])


def validate(mode_rows: list[dict[str, str]], aggregates: list[dict[str, str]],
             winners: list[dict[str, str]]) -> None:
    assert len(mode_rows) == 300, f"expected 300 mode rows, got {len(mode_rows)}"
    assert len(aggregates) == 25, f"expected 25 aggregates, got {len(aggregates)}"
    assert len(winners) == 12, f"expected 12 winners, got {len(winners)}"
    assert len({row["mode"] for row in mode_rows}) == 25
    assert len({row["file"] for row in mode_rows}) == 12
    assert all(row["roundtrip"] == "PASS" for row in mode_rows)
    assert len({row["codec_sha256"] for row in mode_rows}) == 1
    assert sum(as_int(row, "pass_rows") for row in aggregates) == 300


def add_table_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]],
                    numeric_formats: dict[str, str] | None = None) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_index, header in enumerate(headers, 1):
        values = [str(row.get(header, "")) for row in rows]
        width = min(56, max(len(header), *(len(value) for value in values)) + 2)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
        if numeric_formats and header in numeric_formats:
            for cell in sheet[get_column_letter(column_index)][1:]:
                cell.number_format = numeric_formats[header]
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
        if "roundtrip" in headers:
            status_cell = row[headers.index("roundtrip")]
            if status_cell.value == "PASS":
                status_cell.fill = PatternFill("solid", fgColor="E2F0D9")


def write_markdown(aggregates: list[dict[str, str]], winners: list[dict[str, str]],
                   source_rows: list[dict[str, str]], mode_rows: list[dict[str, str]]) -> None:
    lines = [
        "# D40 Detailed Result Table",
        "",
        "Source: `r2-silesia-32k-currenthash-20260821-d40`.",
        "This is a derived view; the source TSV files are unchanged.",
        "",
        f"- Modes: {len(aggregates)} (Auto + 24 forced)",
        f"- Files: {len(winners)} Silesia files, 32 KiB each",
        f"- Rows: {len(mode_rows)}; PASS: {sum(row['roundtrip'] == 'PASS' for row in mode_rows)}",
        f"- Codec SHA-256: `{mode_rows[0]['codec_sha256']}`",
        "",
        "## Mode Summary",
        "",
        "| Mode | Rows | Archive bytes | Ratio | Bits/byte | Encode s | Decode s | Peak RAM MiB | PASS |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in aggregates:
        lines.append(
            f"| {row['mode']} | {row['rows']} | {row['archive_bytes']} | "
            f"{as_float(row, 'ratio'):.6f} | {as_float(row, 'bpb'):.6f} | "
            f"{as_float(row, 'encode_seconds'):.3f} | {as_float(row, 'decode_seconds'):.3f} | "
            f"{as_float(row, 'peak_ram_mib'):.3f} | {row['pass_rows']} |"
        )
    lines += [
        "",
        "## Per-file Winners",
        "",
        "| File | Winner mode(s) | Winner archive bytes | Auto archive bytes | Auto gap bytes |",
        "| --- | --- | ---: | ---: | ---: |",
    ]
    for row in winners:
        lines.append(
            f"| {row['file']} | {row['winner_modes']} | {row['winner_archive_bytes']} | "
            f"{row['auto_archive_bytes']} | {row['auto_gap_bytes']} |"
        )
    lines += [
        "",
        "## Source Packages",
        "",
        "| Mode | Package | Rows | Results CSV SHA-256 |",
        "| --- | --- | ---: | --- |",
    ]
    for row in source_rows:
        lines.append(
            f"| {row['mode']} | `{row['package']}` | {row['rows']} | `{row['results_csv_sha256']}` |"
        )
    lines += [
        "",
        "## Full Per-file Matrix",
        "",
        "The complete 300-row matrix is in `mode_rows.tsv` and the XLSX sheet "
        "`Detailed Rows`; this report keeps the overview compact for review.",
        "",
        "| Mode | File | Archive bytes | Ratio | Encode s | Decode s | Peak RAM MiB | Result | Block type |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | --- | --- |",
    ]
    for row in mode_rows:
        lines.append(
            f"| {row['mode']} | {row['file']} | {row['archive_bytes']} | "
            f"{as_float(row, 'ratio'):.6f} | {as_float(row, 'encode_seconds'):.3f} | "
            f"{as_float(row, 'decode_seconds'):.3f} | {as_float(row, 'peak_ram_mib'):.3f} | "
            f"{row['roundtrip']} | {row['block_types']} |"
        )
    MARKDOWN.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if XLSX.exists() or MARKDOWN.exists():
        raise SystemExit("refusing to overwrite an existing D40 derived table")
    mode_rows = read_tsv(D40 / "mode_rows.tsv")
    aggregates = read_tsv(D40 / "portfolio_aggregate.tsv")
    winners = read_tsv(D40 / "per_file_winners.tsv")
    source_rows = read_tsv(D40 / "source_manifest.tsv")
    validate(mode_rows, aggregates, winners)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["D40 metric", "Value"])
    overview_rows = [
        ("Source directory", str(D40)),
        ("Modes", len(aggregates)),
        ("Files", len(winners)),
        ("Scope KiB", 32),
        ("Normalized rows", len(mode_rows)),
        ("PASS rows", sum(row["roundtrip"] == "PASS" for row in mode_rows)),
        ("Input bytes per mode", as_int(aggregates[0], "input_bytes")),
        ("Codec SHA-256", mode_rows[0]["codec_sha256"]),
        ("Portfolio aggregate SHA-256", sha256(D40 / "portfolio_aggregate.tsv")),
        ("Detailed matrix SHA-256", sha256(D40 / "mode_rows.tsv")),
    ]
    for item in overview_rows:
        overview.append(list(item))
    overview.freeze_panes = "A2"
    overview.column_dimensions["A"].width = 32
    overview.column_dimensions["B"].width = 96
    for cell in overview[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    add_table_sheet(workbook, "Mode Summary", aggregates, {
        "ratio": "0.000000", "bpb": "0.000000", "encode_seconds": "0.000",
        "decode_seconds": "0.000", "peak_ram_mib": "0.000",
    })
    add_table_sheet(workbook, "File Winners", winners)
    add_table_sheet(workbook, "Detailed Rows", mode_rows, {
        "ratio": "0.000000", "bpb": "0.000000", "encode_seconds": "0.000",
        "decode_seconds": "0.000", "peak_ram_mib": "0.000",
    })
    add_table_sheet(workbook, "Sources", source_rows)
    workbook.save(XLSX)
    write_markdown(aggregates, winners, source_rows, mode_rows)
    print(f"wrote {XLSX}")
    print(f"wrote {MARKDOWN}")
    print(f"validated modes={len(aggregates)} rows={len(mode_rows)} pass=300 files={len(winners)}")


if __name__ == "__main__":
    main()
