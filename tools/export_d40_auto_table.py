#!/usr/bin/env python3
"""Export the D40 Auto rows as a standalone exact-value table and figure."""

from __future__ import annotations

import csv
import hashlib
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib.colors import Normalize
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
D40 = ROOT / "results" / "analysis" / "r2-silesia-32k-currenthash-20260821-d40"
OUT = D40 / "auto_only"
MARKDOWN = OUT / "D40_auto_detailed_table.md"
XLSX = OUT / "D40_auto_detailed_table.xlsx"
PNG = OUT / "D40_auto_detailed_table.png"
SVG = OUT / "D40_auto_detailed_table.svg"


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream, delimiter="\t"))


def as_int(row: dict[str, str], key: str) -> int:
    return int(float(row[key]))


def as_float(row: dict[str, str], key: str) -> float:
    return float(row[key])


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def validate(rows: list[dict[str, str]], aggregate: dict[str, str]) -> None:
    if len(rows) != 12:
        raise ValueError(f"expected 12 Auto rows, got {len(rows)}")
    if len({row["file"] for row in rows}) != 12:
        raise ValueError("Auto rows do not contain 12 unique files")
    if any(row["mode"] != "auto" for row in rows):
        raise ValueError("non-Auto row found")
    if any(row["roundtrip"] != "PASS" for row in rows):
        raise ValueError("Auto rows contain a non-PASS result")
    if len({row["codec_sha256"] for row in rows}) != 1:
        raise ValueError("Auto rows use more than one codec hash")
    if as_int(aggregate, "rows") != 12 or as_int(aggregate, "pass_rows") != 12:
        raise ValueError("Auto aggregate row count is not 12/12")
    input_total = sum(as_int(row, "input_bytes") for row in rows)
    archive_total = sum(as_int(row, "archive_bytes") for row in rows)
    if input_total != as_int(aggregate, "input_bytes"):
        raise ValueError("Auto input aggregate mismatch")
    if archive_total != as_int(aggregate, "archive_bytes"):
        raise ValueError("Auto archive aggregate mismatch")
    expected_bpb = 8.0 * archive_total / input_total
    if abs(expected_bpb - as_float(aggregate, "bpb")) > 1e-9:
        raise ValueError("Auto bpb aggregate mismatch")


def write_markdown(rows: list[dict[str, str]], aggregate: dict[str, str]) -> None:
    lines = [
        "# D40 Auto-only Detailed Result Table",
        "",
        "Source: `../mode_rows.tsv`, filtered to `mode=auto`; source files are unchanged.",
        "",
        f"- Files: {len(rows)} Silesia files, 32 KiB each",
        f"- Input bytes: {as_int(aggregate, 'input_bytes')}",
        f"- Archive bytes: {as_int(aggregate, 'archive_bytes')}",
        f"- Aggregate ratio: {as_float(aggregate, 'ratio'):.6f}",
        f"- Aggregate Bpb: {as_float(aggregate, 'bpb'):.6f}",
        f"- PASS: {as_int(aggregate, 'pass_rows')}/{as_int(aggregate, 'rows')}",
        f"- Codec SHA-256: `{rows[0]['codec_sha256']}`",
        "",
        "| File | Selected path | Archive bytes | Ratio | Bpb | Encode s | Decode s | Peak RAM MiB | PASS |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in rows:
        lines.append(
            f"| {row['file']} | {row['block_types'].replace('=1', '')} | "
            f"{as_int(row, 'archive_bytes')} | {as_float(row, 'ratio'):.6f} | "
            f"{as_float(row, 'bpb'):.6f} | {as_float(row, 'encode_seconds'):.6f} | "
            f"{as_float(row, 'decode_seconds'):.6f} | "
            f"{as_float(row, 'peak_ram_mib'):.6f} | {row['roundtrip']} |"
        )
    lines += [
        "",
        "## Aggregate",
        "",
        "| Input bytes | Archive bytes | Ratio | Bpb | Encode s | Decode s | Peak RAM MiB | PASS |",
        "| ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
        f"| {as_int(aggregate, 'input_bytes')} | {as_int(aggregate, 'archive_bytes')} | "
        f"{as_float(aggregate, 'ratio'):.6f} | {as_float(aggregate, 'bpb'):.6f} | "
        f"{as_float(aggregate, 'encode_seconds'):.6f} | {as_float(aggregate, 'decode_seconds'):.6f} | "
        f"{as_float(aggregate, 'peak_ram_mib'):.6f} | "
        f"{as_int(aggregate, 'pass_rows')}/{as_int(aggregate, 'rows')} |",
        "",
        "This is a derived Auto view. It does not claim broader-corpus performance or statistical generalization.",
    ]
    MARKDOWN.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_xlsx(rows: list[dict[str, str]], aggregate: dict[str, str]) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Auto Overview"
    overview.append(["D40 Auto metric", "Value"])
    overview_values = [
        ("Source", str(D40 / "mode_rows.tsv")),
        ("Files", len(rows)),
        ("Scope KiB", 32),
        ("Input bytes", as_int(aggregate, "input_bytes")),
        ("Archive bytes", as_int(aggregate, "archive_bytes")),
        ("Aggregate ratio", as_float(aggregate, "ratio")),
        ("Aggregate Bpb", as_float(aggregate, "bpb")),
        ("Encode seconds", as_float(aggregate, "encode_seconds")),
        ("Decode seconds", as_float(aggregate, "decode_seconds")),
        ("Peak RAM MiB", as_float(aggregate, "peak_ram_mib")),
        ("PASS", f"{as_int(aggregate, 'pass_rows')}/{as_int(aggregate, 'rows')}"),
        ("Codec SHA-256", rows[0]["codec_sha256"]),
    ]
    for item in overview_values:
        overview.append(list(item))
    overview.freeze_panes = "A2"
    overview.column_dimensions["A"].width = 25
    overview.column_dimensions["B"].width = 100
    for cell in overview[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    sheet = workbook.create_sheet("Auto Rows")
    headers = [
        "file", "selected_path", "input_bytes", "archive_bytes", "ratio", "Bpb",
        "encode_seconds", "decode_seconds", "peak_ram_mib", "roundtrip",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row["file"], row["block_types"].replace("=1", ""), as_int(row, "input_bytes"),
            as_int(row, "archive_bytes"), as_float(row, "ratio"), as_float(row, "bpb"),
            as_float(row, "encode_seconds"), as_float(row, "decode_seconds"),
            as_float(row, "peak_ram_mib"), row["roundtrip"],
        ])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 25, 14, 16, 12, 12, 16, 16, 16, 12]
    for index, (header, width) in enumerate(zip(headers, widths), 1):
        sheet.cell(1, index).font = Font(color="FFFFFF", bold=True)
        sheet.cell(1, index).fill = PatternFill("solid", fgColor="1F4E78")
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
        row[-1].fill = PatternFill("solid", fgColor="E2F0D9")
    for column in ("E", "F"):
        for cell in sheet[column][1:]:
            cell.number_format = "0.000000"
    for column in ("G", "H", "I"):
        for cell in sheet[column][1:]:
            cell.number_format = "0.000000"
    workbook.save(XLSX)


def write_png(rows: list[dict[str, str]], aggregate: dict[str, str]) -> None:
    columns = [
        "File", "Selected path", "Archive\nbytes", "Ratio", "Bpb",
        "Encode\ns", "Decode\ns", "Peak RAM\nMiB", "PASS",
    ]
    values = []
    for row in rows:
        values.append([
            row["file"], row["block_types"].replace("=1", ""),
            f"{as_int(row, 'archive_bytes'):,}", f"{as_float(row, 'ratio'):.6f}",
            f"{as_float(row, 'bpb'):.6f}", f"{as_float(row, 'encode_seconds'):.6f}",
            f"{as_float(row, 'decode_seconds'):.6f}",
            f"{as_float(row, 'peak_ram_mib'):.6f}", row["roundtrip"],
        ])
    values.append([
        "TOTAL", "Auto aggregate", f"{as_int(aggregate, 'archive_bytes'):,}",
        f"{as_float(aggregate, 'ratio'):.6f}", f"{as_float(aggregate, 'bpb'):.6f}",
        f"{as_float(aggregate, 'encode_seconds'):.6f}",
        f"{as_float(aggregate, 'decode_seconds'):.6f}",
        f"{as_float(aggregate, 'peak_ram_mib'):.6f}",
        f"{as_int(aggregate, 'pass_rows')}/{as_int(aggregate, 'rows')}",
    ])

    fig, ax = plt.subplots(figsize=(18, 7.2), dpi=240)
    ax.axis("off")
    fig.patch.set_facecolor("white")
    ax.set_title(
        "HybridZip R2 D40 — Auto-only results (32 KiB per file)",
        loc="left", fontsize=18, fontweight="bold", pad=18,
    )
    table = ax.table(
        cellText=values, colLabels=columns, cellLoc="right", colLoc="center",
        loc="upper left", bbox=[0.0, 0.13, 1.0, 0.76],
        colWidths=[0.095, 0.19, 0.105, 0.095, 0.095, 0.11, 0.11, 0.12, 0.08],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(10.2)
    table.scale(1, 1.7)
    bpb_values = [as_float(row, "bpb") for row in rows]
    norm = Normalize(vmin=min(bpb_values), vmax=max(bpb_values))
    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#C7D0D9")
        cell.set_linewidth(0.6)
        cell.PAD = 0.018
        if r == 0:
            cell.set_facecolor("#1F4E78")
            cell.get_text().set_color("white")
            cell.get_text().set_weight("bold")
            cell.get_text().set_ha("center")
        elif r == len(values):
            cell.set_facecolor("#D9EAF7")
            cell.get_text().set_weight("bold")
        else:
            if c == 4:
                cell.set_facecolor(plt.cm.YlGn_r(norm(bpb_values[r - 1])))
            elif c == 8:
                cell.set_facecolor("#E2F0D9")
            elif r % 2 == 0:
                cell.set_facecolor("#F6F8FA")
        if c in (0, 1, 8):
            cell.get_text().set_ha("left" if c != 8 else "center")
    ax.text(
        0.0, 0.075,
        f"Aggregate: 393,216 input B → {as_int(aggregate, 'archive_bytes'):,} archive B | "
        f"ratio {as_float(aggregate, 'ratio'):.6f} | Bpb {as_float(aggregate, 'bpb'):.6f} | "
        f"12/12 PASS | codec {rows[0]['codec_sha256'][:16]}…",
        transform=ax.transAxes, fontsize=10.5, color="#344054",
    )
    ax.text(
        0.0, 0.028,
        "Derived from D40 mode_rows.tsv; no experiment was rerun. Bpb = 8 × archive bytes / input bytes.",
        transform=ax.transAxes, fontsize=9.5, color="#667085",
    )
    fig.savefig(PNG, dpi=240, bbox_inches="tight", facecolor="white")
    fig.savefig(SVG, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    rows = [row for row in read_tsv(D40 / "mode_rows.tsv") if row["mode"] == "auto"]
    aggregate_rows = [
        row for row in read_tsv(D40 / "portfolio_aggregate.tsv") if row["mode"] == "auto"
    ]
    if len(aggregate_rows) != 1:
        raise ValueError("expected exactly one Auto aggregate row")
    aggregate = aggregate_rows[0]
    validate(rows, aggregate)
    write_markdown(rows, aggregate)
    write_xlsx(rows, aggregate)
    write_png(rows, aggregate)
    print(f"wrote {MARKDOWN}")
    print(f"wrote {XLSX}")
    print(f"wrote {PNG}")
    print(f"wrote {SVG}")
    print(f"validated rows={len(rows)} pass={aggregate['pass_rows']} bpb={aggregate['bpb']}")


if __name__ == "__main__":
    main()
